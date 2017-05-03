from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
import argparse

if __name__ == "__main__":

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "input_file",
        help="input image."
    )
    parser.add_argument(
        "output_file",
        help="output xlsx filename."
    )
    parser.add_argument(
        "--mosaic",
        type=int,
        default=4,
        help="how many pixeles per cell."
    )
    parser.add_argument(
        "--max-width",
        type=int,
        default=256
    )

    args = parser.parse_args()
    original_image = Image.open(args.input_file)
    resize_image = original_image.resize(
        (args.max_width, int(float(original_image.size[1]) / original_image.size[0] * args.max_width)),
        Image.ANTIALIAS)

    wb = Workbook()
    ws = wb.active

    image_width = resize_image.size[0]
    image_height = resize_image.size[1]

    for row in range(1, image_height + 1):
        ws.row_dimensions[row].height = args.mosaic
        for col in range(1, image_width + 1):
            ws.column_dimensions[get_column_letter(col)].width = args.mosaic / 5.9
            color = resize_image.getpixel((col - 1, row - 1))
            cell = ws.cell(column=col, row=row)
            cell.fill = PatternFill("solid", fgColor="%02x%02x%02x" % color)

    wb.save(args.output_file if args.output_file.endswith("xlsx") or args.output_file.endswith("xls")
            else args.output_file + ".xlsx")
